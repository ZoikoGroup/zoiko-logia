import io

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.db.base import Base
from app.domains.identity.models import Tenant, User
from app.domains.kriton_workspace import artifacts as artifact_service
from app.domains.kriton_workspace import documents as document_service
from app.domains.kriton_workspace.documents import ingest_document
from app.domains.orchestration_state.schemas import AgentRunCreate
from app.domains.orchestration_state import service as agent_service
from app.domains.orchestration_state.service import (
    classify_agent_goal, create_agent_run, execute_agent_run, serialize_agent_run,
)


def _sales_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Date", "Customer", "Product Category", "Deal Type", "Revenue", "Profit", "Status"])
    sheet.append(["2025-01-03", "Beta Ltd", "Services", "Support", 90000, 30000, "Completed"])
    sheet.append(["2025-01-14", "Apex Corp", "Software", "Annual License", 125000, 83000, "Completed"])
    sheet.append(["2025-02-05", "Delta Inc", "Hardware", "Server", 80000, 20000, "Completed"])
    sheet.append(["2025-02-18", "Core Systems", "Software", "Cloud Subscription", 145000, 97000, "Completed"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_agent_goal_classifies_document_transformation_as_low_risk():
    task_type, risk, plan = classify_agent_goal(
        "Generate an Excel sales report with the highest sale for each month."
    )
    assert task_type == "MONTHLY_HIGHEST_SALES_REPORT"
    assert risk == "LOW"
    assert plan == [
        "inspect_documents",
        "calculate_monthly_highest_sales",
        "create_monthly_sales_xlsx",
    ]


def test_agent_goal_rejects_professional_judgment_in_first_release():
    with pytest.raises(HTTPException) as exc:
        classify_agent_goal(
            "Generate the monthly sales report and advise whether it complies with IFRS."
        )
    assert exc.value.status_code == 422


@pytest.mark.asyncio
async def test_agent_run_generates_verified_monthly_sales_artifact(monkeypatch, tmp_path):
    async def _audit_noop(*args, **kwargs):
        return None

    monkeypatch.setattr(agent_service, "record_event_async", _audit_noop)
    document_root = tmp_path / "data" / "workspace_documents"
    artifact_root = tmp_path / "data" / "workspace_artifacts"
    monkeypatch.setattr(document_service, "_STORAGE_ROOT", document_root)
    monkeypatch.setattr(document_service, "uses_supabase", lambda: False)
    monkeypatch.setattr(artifact_service, "_ROOT", artifact_root)
    monkeypatch.setattr(artifact_service, "uses_supabase", lambda: False)

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as connection:
        await connection.run_sync(Base.metadata.create_all)
    sessions = async_sessionmaker(engine, expire_on_commit=False)
    try:
        async with sessions() as db:
            db.add(Tenant(id="tenant-1", name="Test Tenant"))
            db.add(User(
                id="user-1", tenant_id="tenant-1", email="agent@example.com",
                full_name="Agent Tester", role="Admin",
            ))
            await db.commit()
            document, _ = await ingest_document(
                db,
                filename="sales.xlsx",
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                content=_sales_workbook(),
                tenant_id="tenant-1",
                user_id="user-1",
            )
            run = await create_agent_run(
                db,
                payload=AgentRunCreate(
                    goal="Generate an Excel sales report showing the highest sale for every month.",
                    document_ids=[document.id],
                    conversation_id="conversation-1",
                ),
                tenant_id="tenant-1",
                user_id="user-1",
            )
            run = await execute_agent_run(db, run)
            public = await serialize_agent_run(db, run)

            assert public.status == "COMPLETED"
            assert public.current_step == 3
            assert [step.status for step in public.steps] == ["COMPLETED"] * 3
            assert public.final_response["months"] == 2
            assert public.final_response["selected_revenue_total"] == 270000
            assert len(public.artifacts) == 1

            stored = next(artifact_root.rglob("*.xlsx"))
            workbook = load_workbook(stored, read_only=True, data_only=True)
            try:
                assert "Monthly Highest Sales" in workbook.sheetnames
                rows = list(workbook["Monthly Highest Sales"].iter_rows(values_only=True))
                assert rows[1][0] == "2025-01"
                assert 125000 in rows[1]
                assert rows[2][0] == "2025-02"
                assert 145000 in rows[2]
            finally:
                workbook.close()

            # Completed runs are idempotent and do not create another artifact.
            resumed = await execute_agent_run(db, run)
            resumed_public = await serialize_agent_run(db, resumed)
            assert resumed_public.status == "COMPLETED"
            assert len(resumed_public.artifacts) == 1
    finally:
        await engine.dispose()
